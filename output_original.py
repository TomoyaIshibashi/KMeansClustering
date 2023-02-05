# -*- coding : UTF-8 -*-

# .ffs(F0データ)を行列化するためExcelファイルへ出力

import os
import glob
import pandas as pd
import openpyxl

wb1 = openpyxl.load_workbook('/Users/6ashi/Documents/MATLAB/EXL/EXL_N/F0data_N_select_1.xlsx')
ws1 = wb1['Sheet1']

wb2 = openpyxl.load_workbook('/Users/6ashi/Documents/MATLAB/EXL/EXL_N/F0data_N_select_2.xlsx')
ws2 = wb2['Sheet1']

wb3 = openpyxl.load_workbook('/Users/6ashi/Documents/MATLAB/EXL/EXL_N/F0data_N_select_3.xlsx')
ws3 = wb3['Sheet1']

wb4 = openpyxl.load_workbook('/Users/6ashi/Documents/MATLAB/EXL/EXL_N/F0data_N_select_4.xlsx')
ws4 = wb4['Sheet1']

# F0データ(smoothed)のフォルダのパスを取得
DirectorySession_lc_original = '/Users/6ashi/Documents/MATLAB/balance_F02_LC/balance_F02_LC_N/FFS'

# F0データ(smoothed)のフォルダのパスを取得
DirectorySession_fc_original = '/Users/6ashi/Documents/MATLAB/balance_F02_FC/balance_F02_FC_N/FFS'

# F0データ(smoothed)のフォルダのパスを取得
DirectorySession_lc_target = '/Users/6ashi/Documents/MATLAB/balance_F02_LC/balance_F02_LC_A/FFS'

# F0データ(smoothed)のフォルダのパスを取得
DirectorySession_fc_target = '/Users/6ashi/Documents/MATLAB/balance_F02_FC/balance_F02_FC_A/FFS'

# F0データファイルを読み込んで行列化
os.chdir(DirectorySession_lc_original) # 原音声の Last Chunk
fileList = sorted(glob.glob('*.ffs'))
cnt1 = cnt2 = cnt3 = cnt4 = 0
for file_lc in fileList:
  fnCompos_lc = file_lc.split('_')
  # balance_F02_LC_A_0001.ffs
  fileId_lc = fnCompos_lc[0] + '_' + fnCompos_lc[1] + '_' + fnCompos_lc[2] + '_' + fnCompos_lc[3] + '_' + fnCompos_lc[4]
  fileId_lc_target = fnCompos_lc[0] + '_' + fnCompos_lc[1] + '_' + fnCompos_lc[2] + '_A_' + fnCompos_lc[4]
  os.chdir(DirectorySession_lc_original)
  f_lc = open(file_lc, 'r')
  tsdata_lc = f_lc.readlines()
  f_lc.close
  # 長さ（配列数）が150-260のものを抽出
  if 150 <= len(tsdata_lc) and len(tsdata_lc) < 260:
    os.chdir(DirectorySession_lc_target)
    if os.path.exists(fileId_lc_target):
      g = open(fileId_lc_target, 'r')
      tsdata_lc_target = g.readlines()
      g.close
      if 1.5 * len(tsdata_lc) > len(tsdata_lc_target):
        cnt1 += 1 
        # ファイルの名称
        ws1.cell(1,cnt1,value = fileId_lc)
        #print(fnCompos[4])
        # Excelへ書き込み
        for i in range(0, len(tsdata_lc)):
          ws1.cell(i+2,cnt1,value = tsdata_lc[i])
        wb1.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_N/F0data_N_select_1.xlsx')
        #for i in range(0, len(tsdata_lc_target)):
        #  ws1.cell(i+2,cnt1,value = tsdata_lc_target[i])
        #wb1.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_A/F0data_A_select_1.xlsx')
  
  # 長さ（配列数）が260-370のものを抽出
  elif 260 <= len(tsdata_lc) and len(tsdata_lc) < 370:
    os.chdir(DirectorySession_lc_target)
    if os.path.exists(fileId_lc_target):
      g = open(fileId_lc_target, 'r')
      tsdata_target = g.readlines()
      g.close
      if 1.5 * len(tsdata_lc) > len(tsdata_lc_target):
        cnt2 += 1
        # ファイルの名称
        ws2.cell(1,cnt2,value = fileId_lc)
        #print(fnCompos[4])
        # Excelへ書き込み
        for i in range(0, len(tsdata_lc)):
          ws2.cell(i+2,cnt2,value = tsdata_lc[i])
        wb2.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_N/F0data_N_select_2.xlsx')
        #for i in range(0, len(tsdata_lc_target)):
        #  ws2.cell(i+2,cnt2,value = tsdata_lc_target[i])
        #wb2.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_A/F0data_A_select_2.xlsx')

  # 長さ（配列数）が370-480のものを抽出
  elif 370 <= len(tsdata_lc) and len(tsdata_lc) < 480:
    os.chdir(DirectorySession_lc_target)
    if os.path.exists(fileId_lc_target):
      g = open(fileId_lc_target, 'r')
      tsdata_lc_target = g.readlines()
      g.close
      if 1.5 * len(tsdata_lc) > len(tsdata_lc_target):
        cnt3 += 1
        # ファイルの名称
        ws3.cell(1,cnt3,value = fileId_lc)
        #print(fnCompos[4])
        # Excelへ書き込み
        for i in range(0, len(tsdata_lc)):
          ws3.cell(i+2,cnt3,value = tsdata_lc[i])
        wb3.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_N/F0data_N_select_3.xlsx')
        #for i in range(0, len(tsdata_lc_target)):
        #  ws3.cell(i+2,cnt3,value = tsdata_lc_target[i])
        #wb3.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_A/F0data_A_select_3.xlsx')

  # 長さ（配列数）が480-590のものを抽出
  elif 480 <= len(tsdata_lc) and len(tsdata_lc) < 590:
    os.chdir(DirectorySession_lc_target)
    if os.path.exists(fileId_lc_target):
      g = open(fileId_lc_target, 'r')
      tsdata_lc_target = g.readlines()
      g.close
      if 1.5 * len(tsdata_lc) > len(tsdata_lc_target):
        cnt4 += 1
        # ファイルの名称
        ws4.cell(1,cnt4,value = fileId_lc)
        #print(fnCompos[4])
        # Excelへ書き込み
        for i in range(0, len(tsdata_lc)):
          ws4.cell(i+2,cnt4,value = tsdata_lc[i])
        wb4.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_N/F0data_N_select_4.xlsx')
        #for i in range(0, len(tsdata_lc_target)):
        #  ws4.cell(i+2,cnt4,value = tsdata_lc_target[i])
        #wb4.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_A/F0data_A_select_4.xlsx')


os.chdir(DirectorySession_fc_original)
fileList = sorted(glob.glob('*.ffs'))
for file_fc in fileList:
  fnCompos_fc = file_fc.split('_')
  # balance_F02_LC_A_0001.ffs
  fileId_fc = fnCompos_fc[0] + '_' + fnCompos_fc[1] + '_' + fnCompos_fc[2] + '_' + fnCompos_fc[3] + '_' + fnCompos_fc[4]
  fileId_fc_target = fnCompos_fc[0] + '_' + fnCompos_fc[1] + '_' + fnCompos_fc[2] + '_A_' + fnCompos_fc[4]
  os.chdir(DirectorySession_fc_original)
  f_fc = open(file_fc, 'r')
  tsdata_fc = f_fc.readlines()
  f_fc.close

  # 長さ（配列数）が150-260のものを抽出
  if 150 <= len(tsdata_fc) and len(tsdata_fc) < 260:
    os.chdir(DirectorySession_fc_target)
    if os.path.exists(fileId_fc_target):
      g = open(fileId_fc_target, 'r')
      tsdata_fc_target = g.readlines()
      g.close
      if 1.5 * len(tsdata_fc) > len(tsdata_fc_target):
        cnt1 += 1
        # ファイルの名称
        ws1.cell(1,cnt1,value = fileId_fc)
        #print(fnCompos[4])
        # Excelへ書き込み
        for i in range(0, len(tsdata_fc)):
          ws1.cell(i+2,cnt1,value = tsdata_fc[i])
        wb1.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_N/F0data_N_select_1.xlsx')
        #for i in range(0, len(tsdata_fc_target)):
        #  ws1.cell(i+2,cnt1,value = tsdata_fc_target[i])
        #wb1.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_A/F0data_A_select_1.xlsx')

  # 長さ（配列数）が260-370のものを抽出
  elif 260 <= len(tsdata_fc) and len(tsdata_fc) < 370:
    os.chdir(DirectorySession_fc_target)
    if os.path.exists(fileId_fc_target):
      g = open(fileId_fc_target, 'r')
      tsdata_fc_target = g.readlines()
      g.close
      if 1.5 * len(tsdata_fc) > len(tsdata_fc_target):
        cnt2 += 1
        # ファイルの名称
        ws2.cell(1,cnt2,value = fileId_fc)
        #print(fnCompos[4])
        # Excelへ書き込み
        for i in range(0, len(tsdata_fc)):
          ws2.cell(i+2,cnt2,value = tsdata_fc[i])
        wb2.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_N/F0data_N_select_2.xlsx')
        #for i in range(0, len(tsdata_fc_target)):
        #  ws2.cell(i+2,cnt2,value = tsdata_fc_target[i])
        #wb2.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_A/F0data_A_select_2.xlsx')

  # 長さ（配列数）が370-480のものを抽出
  if 370 <= len(tsdata_fc) and len(tsdata_fc) < 480: ##　elif → if
    os.chdir(DirectorySession_fc_target)
    if os.path.exists(fileId_fc_target):
      g = open(fileId_fc_target, 'r')
      tsdata_fc_target = g.readlines()
      g.close
      if 1.5 * len(tsdata_fc) > len(tsdata_fc_target):
        cnt3 += 1
        # ファイルの名称
        ws3.cell(1,cnt3,value = fileId_fc)
        #print(fnCompos[4])
        # Excelへ書き込み
        for i in range(0, len(tsdata_fc)):
          ws3.cell(i+2,cnt3,value = tsdata_fc[i])
        wb3.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_N/F0data_N_select_3.xlsx')
        #for i in range(0, len(tsdata_fc_target)):
        #  ws3.cell(i+2,cnt3,value = tsdata_fc_target[i])
        #wb3.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_A/F0data_A_select_3.xlsx')

  # 長さ（配列数）が480-590のものを抽出
  elif 480 <= len(tsdata_fc) and len(tsdata_fc) < 590:
    os.chdir(DirectorySession_fc_target)
    if os.path.exists(fileId_fc_target):
      g = open(fileId_fc_target, 'r')
      tsdata_fc_target = g.readlines()
      g.close
      if 1.5 * len(tsdata_fc) > len(tsdata_fc_target):
        cnt4 += 1
        # ファイルの名称
        ws4.cell(1,cnt4,value = fileId_fc)
        #print(fnCompos[4])
        # Excelへ書き込み
        for i in range(0, len(tsdata_fc)):
          ws4.cell(i+2,cnt4,value = tsdata_fc[i])
        wb4.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_N/F0data_N_select_4.xlsx')
        #for i in range(0, len(tsdata_fc_target)):
        #  ws4.cell(i+2,cnt4,value = tsdata_fc_target[i])
        #wb4.save('/Users/6ashi/Documents/MATLAB/EXL/EXL_A/F0data_A_select_4.xlsx')


