# -*- coding : UTF-8 -*-

# .txt(smoothed & standardized & resampled F0データ)を行列化するためExcelファイルへ出力

import os
import glob
import pandas as pd
import openpyxl

# クラスタ番号
cluster = '1'
group = '1'

# Directoryを取得
directory_in = '/Users/6ashi/Documents/MATLAB/Clustering/Clustering_F02/Clustering_N_A/G' + group
directory_out = '/Users/6ashi/Documents/MATLAB/EXL/EXL_N_A/F0data_N_' + group + '_cluster' + cluster + '.xlsx'
fname_exl = 'F0data_N_' + group + '_cluster' + cluster + '.xlsx'
fname_centroid = 'N_centroid_G' + group + '_' + cluster + '.txt'

# F0データ(smoothed & standardized)のフォルダのパスを取得
DirectorySession_cluster = directory_in + '/cluster' + cluster

# F0データ(cluster's centroid)のフォルダのパスを取得
DirectorySession_centroid = directory_in + '/centroid'

# F0データファイルを読み込んで行列化
os.chdir(DirectorySession_centroid)
f = open(fname_centroid, 'r')
tsdata = f.readlines()
f.close

# Excelへ書き込み
wb_original = openpyxl.load_workbook(directory_out)
ws_original = wb_original['Sheet1']
os.chdir(DirectorySession_cluster)
ws_original.cell(1,1).value = 'centroid'
for i in range(len(tsdata)):
  ws_original.cell(1,i+2).value = tsdata[i]
wb_original.save(directory_out)

fileList = sorted(glob.glob('*.txt'))
cnt = 1
for file in fileList:
  fnCompos = file.split('_')
  # balance_F02_LC_N_0001.sfs
  fileId = fnCompos[0] + '_' + fnCompos[1] + '_' + fnCompos[2] + '_' + fnCompos[3] + '_' + fnCompos[4]

  os.chdir(DirectorySession_cluster)
  f = open(file, 'r')
  tsdata = f.readlines()
  f.close

  cnt += 1

  # Excelへ書き込み
  wb_original = openpyxl.load_workbook(directory_out)
  ws_original = wb_original['Sheet1']
  os.chdir(DirectorySession_cluster)
  ws_original.cell(cnt,1).value = fileId
  for i in range(len(tsdata)):
    ws_original.cell(cnt,i+2).value = str(tsdata[i])
  wb_original.save(directory_out)

